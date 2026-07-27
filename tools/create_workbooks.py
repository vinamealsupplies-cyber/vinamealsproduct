from __future__ import annotations

from copy import copy
from pathlib import Path
import csv

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "import-templates"
# Thư mục public thật của app Next (trước đây trỏ nhầm vào "starter/public",
# nên bản template tải về từ web không bao giờ được cập nhật).
PUBLIC = ROOT / "public" / "templates"
OUT.mkdir(parents=True, exist_ok=True)
PUBLIC.mkdir(parents=True, exist_ok=True)

INK = "18312C"
GREEN = "0B7A55"
GREEN_DARK = "075B40"
MINT = "DFF7E9"
LIME = "D8F27A"
YELLOW = "FFD76A"
ORANGE = "FF875E"
BLUE = "BCE7F5"
LAVENDER = "E7D8FF"
PINK = "FFD9DF"
WHITE = "FFFFFF"
LIGHT = "F4F8F6"
LINE = "D5E3DC"
MUTED = "60736E"
RED = "FDE7E4"

HEADERS = [
    "operation", "product_handle", "product_name", "slug", "short_description", "description",
    "category_path", "variant_name", "sku", "barcode", "attributes_json",
    "retail_price", "sale_price", "wholesale_price", "cost_price",
    "track_inventory", "opening_quantity", "location_code",
    "taxable", "unit", "weight_oz", "status", "featured",
    *[f"image_url_{index}" for index in range(1, 11)], "video_url"
]
REQUIRED = {"operation", "product_handle", "product_name", "sku", "retail_price", "cost_price"}
HEADER_GROUPS = {
    "operation": BLUE,
    "product_handle": MINT, "product_name": MINT, "slug": MINT, "short_description": MINT,
    "description": MINT, "category_path": MINT,
    "variant_name": LAVENDER, "sku": LAVENDER, "barcode": LAVENDER, "attributes_json": LAVENDER,
    "retail_price": YELLOW, "sale_price": YELLOW, "wholesale_price": YELLOW, "cost_price": YELLOW,
    "track_inventory": PINK, "opening_quantity": PINK, "location_code": PINK,
    "taxable": ORANGE, "unit": ORANGE, "weight_oz": ORANGE, "status": ORANGE, "featured": ORANGE,
    **{f"image_url_{index}": BLUE for index in range(1, 11)}, "video_url": BLUE,
}
DESCRIPTIONS = {
    "operation": "CREATE, UPDATE, or UPSERT. UPSERT is the safest default for a reusable file.",
    "product_handle": "Groups variants into one product. Repeat the same handle for multiple SKUs.",
    "product_name": "Customer-facing English product name.",
    "slug": "Optional lowercase URL slug using letters, numbers, and hyphens.",
    "short_description": "Recommended maximum 180 characters for catalog cards.",
    "description": "Full English description. Include package, storage, ingredients, and serving information as applicable.",
    "category_path": "Category hierarchy such as Snacks > Fruit. The importer may create missing categories only when that option is explicitly enabled.",
    "variant_name": "Sellable option such as 8 oz bag or 20 count.",
    "sku": "Unique stock keeping unit. Case-insensitive uniqueness is enforced.",
    "barcode": "Optional UPC/EAN/GTIN. Must be unique when present.",
    "attributes_json": 'Optional JSON object, for example {"size":"8 oz","flavor":"mango"}.',
    "retail_price": "Retail (list) price in USD, 0 or greater.",
    "sale_price": "Optional promotional price. Leave blank for no sale. When set, must be lower than retail_price; storefront strikes through retail.",
    "wholesale_price": "Optional wholesale price in USD. Wholesale pricing does not automatically make the customer tax exempt.",
    "cost_price": "Unit cost / giá nhập in USD, 0 or greater.",
    "track_inventory": "TRUE or FALSE.",
    "opening_quantity": "Opening on-hand quantity. Use only when creating a new SKU; later changes should use inventory movements.",
    "location_code": "Inventory location code, for example MAIN.",
    "taxable": "TRUE or FALSE. Final tax treatment also depends on customer and jurisdiction.",
    "unit": "Each, bag, box, case, bottle, jar, pack, pound, ounce, or custom unit.",
    "weight_oz": "Optional shipping weight in ounces.",
    "status": "draft, active, or archived. Prefer this over legacy active TRUE/FALSE.",
    "featured": "TRUE or FALSE. Featured products appear first in the catalog.",
    "video_url": "Optional existing video URL. For new uploads, use Cloudflare Stream through the admin UI.",
}
for index in range(1, 11):
    DESCRIPTIONS[f"image_url_{index}"] = f"Optional image URL {index}. A product may have no more than 10 images. image_url_1 becomes the cover image."

# Columns: op, handle, name, slug, short, desc, cat, variant, sku, barcode, attrs,
# retail, sale, wholesale, cost, track, opening, location, taxable, unit, weight, status, featured, images×10, video
SAMPLE_ROWS = [
    ["UPSERT", "tropical-mango-slices", "Tropical Mango Slices", "tropical-mango-slices", "Sweet, sunny mango slices ready to enjoy.", "Bright tropical mango slices with a soft bite and naturally sweet flavor.", "Snacks > Fruit", "8 oz bag", "MANGO-8OZ", "", '{"size":"8 oz"}', 8.99, 6.99, 6.25, 3.40, True, 42, "MAIN", True, "bag", 8, "active", True, "https://media.example.com/products/mango-1.webp", "https://media.example.com/products/mango-2.webp", "", "", "", "", "", "", "", "", ""],
    ["UPSERT", "garden-veggie-dumplings", "Garden Veggie Dumplings", "garden-veggie-dumplings", "Tender dumplings filled with colorful vegetables.", "Freezer-friendly vegetable dumplings for quick lunches and family meals.", "Frozen > Dumplings", "20 count", "DUMP-VEG-20", "", '{"count":20}', 12.99, "", 9.40, 5.20, True, 18, "MAIN", True, "bag", 24, "active", True, "https://media.example.com/products/dumplings-1.webp", "", "", "", "", "", "", "", "", "", ""],
    ["UPSERT", "golden-chili-crisp", "Golden Chili Crisp", "golden-chili-crisp", "Crunchy, savory, and gently spicy.", "A spoonable chili crisp for noodles, rice, eggs, and vegetables.", "Sauces > Chili", "6 oz jar", "CHILI-6OZ", "", '{"size":"6 oz"}', 10.50, 8.50, 7.60, 3.85, True, 27, "MAIN", True, "jar", 11, "active", False, "https://media.example.com/products/chili-1.webp", "", "", "", "", "", "", "", "", "", ""],
    ["UPSERT", "pure-coconut-water", "Pure Coconut Water", "pure-coconut-water", "Clean, refreshing coconut water.", "Serve chilled for a crisp and refreshing drink.", "Beverages > Coconut Water", "16.9 fl oz bottle", "COCO-169", "", '{"volume":"16.9 fl oz"}', 3.99, "", 2.65, 1.35, True, 64, "MAIN", True, "bottle", 18, "draft", False, "https://media.example.com/products/coconut-1.webp", "", "", "", "", "", "", "", "", "", ""],
]


def title_style(cell, fill=GREEN_DARK, size=16):
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(color=WHITE, bold=True, size=size)
    cell.alignment = Alignment(vertical="center")


def create_import_template(path: Path) -> None:
    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    products = wb.create_sheet("Products")
    lists = wb.create_sheet("Lists")

    instructions.sheet_view.showGridLines = False
    instructions.merge_cells("A1:H2")
    instructions["A1"] = "Vinameals — Product Import Template"
    title_style(instructions["A1"], size=20)
    instructions["A1"].alignment = Alignment(vertical="center", horizontal="left")
    instructions.row_dimensions[1].height = 32
    instructions.row_dimensions[2].height = 18
    instructions["A4"] = "Purpose"
    instructions["A5"] = "Prepare products, variants, pricing, opening inventory, and media references for a validated admin import. Customer-facing text should be written in English."
    instructions.merge_cells("A5:H6")
    instructions["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    instructions["A5"].font = Font(color=MUTED, size=11)

    steps = [
        ("1", "Keep the header row unchanged. Required: operation, product_handle, product_name, sku, retail_price, cost_price."),
        ("2", "Use one row per SKU. Repeat product_handle and product_name when a product has multiple variants."),
        ("3", "Use UPSERT unless you intentionally need CREATE-only or UPDATE-only behavior."),
        ("4", "sale_price is optional. When set it must be lower than retail_price (storefront shows sale + struck retail)."),
        ("5", "status is draft, active, or archived. featured is TRUE/FALSE. Provide up to 10 image URLs; image_url_1 is the cover."),
        ("6", "Run Preview Import in Admin. Fix every error, review warnings, then commit the validated batch."),
        ("7", "Do not use opening_quantity to correct live stock — use Inventory adjustments. Wholesale price ≠ tax exempt."),
    ]
    instructions["A8"] = "Workflow"
    title_style(instructions["A8"], fill=INK, size=12)
    instructions.merge_cells("A8:H8")
    for row_idx, (number, text) in enumerate(steps, start=9):
        instructions[f"A{row_idx}"] = number
        instructions[f"A{row_idx}"].font = Font(color=GREEN, bold=True, size=12)
        instructions[f"B{row_idx}"] = text
        instructions.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
        instructions[f"B{row_idx}"].alignment = Alignment(wrap_text=True, vertical="top")
        instructions[f"B{row_idx}"].font = Font(color=INK, size=10)
        instructions.row_dimensions[row_idx].height = 31

    instructions["A18"] = "Color guide"
    title_style(instructions["A18"], fill=INK, size=12)
    instructions.merge_cells("A18:H18")
    legends = [("Identity", MINT), ("Variant", LAVENDER), ("Price", YELLOW), ("Inventory", PINK), ("Tax / status", ORANGE), ("Media", BLUE)]
    for col_idx, (label, color) in enumerate(legends, start=1):
        cell = instructions.cell(row=19, column=col_idx)
        cell.value = label
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(color=INK, bold=True, size=9)
        cell.alignment = Alignment(horizontal="center")

    instructions["A22"] = "Allowed values"
    title_style(instructions["A22"], fill=INK, size=12)
    instructions.merge_cells("A22:H22")
    allowed = [
        ("operation", "CREATE, UPDATE, UPSERT"),
        ("status", "draft, active, archived"),
        ("Boolean fields", "TRUE, FALSE (track_inventory, taxable, featured)"),
        ("unit", "each, bag, box, case, bottle, jar, pack, pound, ounce"),
        ("category_path", "Parent > Child > Grandchild"),
        ("attributes_json", '{"key":"value"}'),
        ("sale_price", "Blank or number strictly less than retail_price"),
    ]
    for row_idx, (field, values) in enumerate(allowed, start=23):
        instructions[f"A{row_idx}"] = field
        instructions[f"A{row_idx}"].font = Font(bold=True, color=GREEN_DARK)
        instructions[f"B{row_idx}"] = values
        instructions.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
        instructions[f"B{row_idx}"].font = Font(color=MUTED)

    # Mô tả từng cột — thay cho chú thích ô ở hàng tiêu đề sheet Products.
    guide_start = 23 + len(allowed) + 2
    instructions[f"A{guide_start - 1}"] = "Field guide"
    title_style(instructions[f"A{guide_start - 1}"], fill=INK, size=12)
    instructions.merge_cells(f"A{guide_start - 1}:H{guide_start - 1}")
    for offset, header in enumerate(HEADERS):
        row_idx = guide_start + offset
        instructions[f"A{row_idx}"] = header
        instructions[f"A{row_idx}"].font = Font(bold=True, color=GREEN_DARK, size=9)
        instructions[f"B{row_idx}"] = DESCRIPTIONS.get(header, "Import field.")
        instructions.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
        instructions[f"B{row_idx}"].alignment = Alignment(wrap_text=True, vertical="top")
        instructions[f"B{row_idx}"].font = Font(color=MUTED, size=9)

    for col in range(1, 9):
        instructions.column_dimensions[chr(64 + col)].width = 18 if col == 1 else 15
    instructions.freeze_panes = "A4"

    products.sheet_view.showGridLines = False
    products.freeze_panes = "A2"
    products.auto_filter.ref = f"A1:{products.cell(1, len(HEADERS)).coordinate}"
    products.row_dimensions[1].height = 52
    thin_bottom = Side(style="thin", color=LINE)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = products.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = PatternFill("solid", fgColor=HEADER_GROUPS[header])
        cell.font = Font(color=INK, bold=True, size=9)
        cell.alignment = Alignment(text_rotation=0, wrap_text=True, vertical="center")
        cell.border = Border(bottom=thin_bottom)
        # KHÔNG dùng cell.comment: openpyxl ghi chú thích ra
        # `xl/comments/comment1.xml`, còn exceljs (thư viện đọc file của app)
        # chỉ nhận `xl/commentsN.xml` theo cách Excel ghi. Lệch quy ước này làm
        # exceljs ném "Cannot read properties of undefined (reading 'comments')"
        # và toàn bộ chức năng import không đọc nổi chính file mẫu.
        # Mô tả từng cột được đưa xuống bảng "Field guide" ở sheet Instructions.
    for row_idx, values in enumerate(SAMPLE_ROWS, start=2):
        for col_idx, value in enumerate(values, start=1):
            cell = products.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(color="0000FF", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=col_idx in (5, 6, 11))
            if HEADERS[col_idx - 1] in {"retail_price", "sale_price", "wholesale_price", "cost_price"}:
                cell.number_format = '$#,##0.00;[Red]($#,##0.00);-'
            elif HEADERS[col_idx - 1] in {"opening_quantity", "weight_oz"}:
                cell.number_format = '#,##0.00;[Red](#,##0.00);-'
    products.row_dimensions[2].height = 42
    products.row_dimensions[3].height = 42
    products.row_dimensions[4].height = 42
    products.row_dimensions[5].height = 42

    widths = {
        "operation": 12, "product_handle": 25, "product_name": 28, "slug": 27,
        "short_description": 38, "description": 55, "category_path": 27,
        "variant_name": 21, "sku": 19, "barcode": 18, "attributes_json": 30,
        "retail_price": 14, "sale_price": 12, "wholesale_price": 16, "cost_price": 14,
        "track_inventory": 16, "opening_quantity": 16,
        "location_code": 14, "taxable": 11, "unit": 12, "weight_oz": 12, "status": 12, "featured": 11,
        **{f"image_url_{index}": 34 for index in range(1, 11)}, "video_url": 34,
    }
    for idx, header in enumerate(HEADERS, start=1):
        products.column_dimensions[products.cell(1, idx).column_letter].width = widths[header]

    # Lists and data validations.
    lists_data = {
        "A": ["Operations", "CREATE", "UPDATE", "UPSERT"],
        "B": ["Booleans", "TRUE", "FALSE"],
        "C": ["Units", "each", "bag", "box", "case", "bottle", "jar", "pack", "pound", "ounce"],
        "D": ["Locations", "MAIN", "FREEZER", "BACKROOM"],
        "E": ["Statuses", "draft", "active", "archived"],
    }
    for column, values in lists_data.items():
        for row, value in enumerate(values, start=1):
            lists[f"{column}{row}"] = value
    lists.sheet_state = "hidden"

    header_index = {header: idx + 1 for idx, header in enumerate(HEADERS)}
    validation_specs = [
        ("operation", "'Lists'!$A$2:$A$4"),
        ("track_inventory", "'Lists'!$B$2:$B$3"),
        ("taxable", "'Lists'!$B$2:$B$3"),
        ("featured", "'Lists'!$B$2:$B$3"),
        ("unit", "'Lists'!$C$2:$C$10"),
        ("location_code", "'Lists'!$D$2:$D$4"),
        ("status", "'Lists'!$E$2:$E$4"),
    ]
    for header, formula in validation_specs:
        dv = DataValidation(type="list", formula1=formula, allow_blank=header not in {"operation"})
        dv.error = "Choose a value from the dropdown."
        dv.errorTitle = "Invalid value"
        dv.prompt = DESCRIPTIONS[header]
        dv.promptTitle = header
        products.add_data_validation(dv)
        col_letter = products.cell(1, header_index[header]).column_letter
        dv.add(f"{col_letter}2:{col_letter}2000")

    decimal_validation = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    decimal_validation.error = "Enter a number greater than or equal to 0."
    products.add_data_validation(decimal_validation)
    for header in ("retail_price", "sale_price", "wholesale_price", "cost_price", "opening_quantity", "weight_oz"):
        col_letter = products.cell(1, header_index[header]).column_letter
        decimal_validation.add(f"{col_letter}2:{col_letter}2000")

    # Visual warnings for missing required values on populated rows.
    red_fill = PatternFill("solid", fgColor=RED)
    for header in REQUIRED:
        col_letter = products.cell(1, header_index[header]).column_letter
        products.conditional_formatting.add(
            f"{col_letter}2:{col_letter}2000",
            FormulaRule(formula=[f'AND($B2<>"",{col_letter}2="")'], fill=red_fill)
        )

    wb.save(path)


def create_backlog(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Backlog"
    ws.sheet_view.showGridLines = False
    headers = ["ID", "Phase", "Workstream", "Task", "Deliverable / acceptance", "Priority", "Owner", "Dependency", "Estimate", "Status"]
    rows = [
        ["FND-001", "0 — Foundation", "Repository", "Create GitHub repository and branch protection", "Main branch protected; PR checks required", "P0", "Tech lead", "", "0.5 day", "Not started"],
        ["FND-002", "0 — Foundation", "Environment", "Create dev/staging/prod environment plan", "Secrets and URLs separated by environment", "P0", "Tech lead", "FND-001", "1 day", "Not started"],
        ["DB-001", "1 — Data", "Supabase", "Apply schema migrations", "All tables, enums, constraints, indexes, triggers created", "P0", "Backend", "FND-002", "1 day", "Not started"],
        ["DB-002", "1 — Data", "Security", "Test row-level security by role", "Anonymous, customer, staff, manager, admin tests pass", "P0", "Backend", "DB-001", "2 days", "Not started"],
        ["AUTH-001", "2 — Accounts", "Authentication", "Connect sign-up, sign-in, email confirmation, sign-out", "Customer account works end to end", "P0", "Full stack", "DB-001", "2 days", "Not started"],
        ["AUTH-002", "2 — Accounts", "Authorization", "Admin navigation and route guards", "Only staff roles can access /admin", "P0", "Full stack", "DB-002,AUTH-001", "1 day", "Not started"],
        ["CAT-001", "3 — Catalog", "Products", "Persist product and variant create/edit", "Product, SKU, price, cost, tax, status persist", "P0", "Full stack", "AUTH-002", "3 days", "Not started"],
        ["CAT-002", "3 — Catalog", "Categories", "Persist category tree and sort order", "Storefront dropdown reflects admin changes", "P0", "Full stack", "CAT-001", "2 days", "Not started"],
        ["MED-001", "3 — Catalog", "Cloudflare R2", "Complete signed image upload flow", "JPEG/PNG/WebP/AVIF uploaded; max 10 enforced", "P0", "Backend", "CAT-001", "2 days", "Not started"],
        ["MED-002", "3 — Catalog", "Cloudflare Stream", "Direct video upload and webhook status", "Video ready state and playback URL synchronized", "P1", "Backend", "CAT-001", "3 days", "Not started"],
        ["INV-001", "4 — Inventory", "Inventory ledger", "Implement receiving, adjustment, reservation, sale, return", "Every balance change has an immutable movement", "P0", "Backend", "DB-002,CAT-001", "4 days", "Not started"],
        ["INV-002", "4 — Inventory", "Admin UI", "Inventory detail/search/sort/filter/export", "Search by name/SKU/category; category and location filters", "P0", "Full stack", "INV-001", "3 days", "Not started"],
        ["IMP-001", "4 — Inventory", "Excel import", "Implement preview and transactional commit", "Valid rows commit atomically; errors do not partially write", "P0", "Backend", "CAT-001,INV-001", "4 days", "Not started"],
        ["CUS-001", "5 — Customers", "Customer master", "Guest, retail, wholesale profiles", "Searchable profiles with price level and balances", "P0", "Full stack", "AUTH-001", "3 days", "Not started"],
        ["TAX-001", "5 — Customers", "Exemption workflow", "Document metadata and approval audit", "Tax exemption cannot be self-approved by customer", "P0", "Backend", "CUS-001", "3 days", "Not started"],
        ["SAL-001", "6 — Sales", "Orders", "Create admin/storefront sales order workflow", "Line snapshots, totals, status transitions, stock reservations", "P0", "Full stack", "INV-001,CUS-001", "5 days", "Not started"],
        ["SAL-002", "6 — Sales", "Invoices", "Issue invoice and PDF/print view", "Sequential number, line snapshots, tax, balance", "P0", "Full stack", "SAL-001", "4 days", "Not started"],
        ["PAY-001", "7 — Payments", "Payments", "Select provider and implement checkout/webhooks", "Idempotent payment and refund reconciliation", "P1", "Full stack", "SAL-002", "6 days", "Deferred"],
        ["REP-001", "8 — Reporting", "Financial reports", "Monthly/yearly sales, received, COGS, expenses, profit", "Definitions documented and totals reconcile", "P0", "Backend", "SAL-002,INV-001", "4 days", "Not started"],
        ["QA-001", "9 — Launch", "Quality", "Accessibility, responsive, security, performance tests", "Acceptance checklist completed", "P0", "QA", "All", "5 days", "Not started"],
        ["OPS-001", "9 — Launch", "Operations", "Backups, monitoring, alerts, incident runbook", "Restore test and owner alerts verified", "P0", "DevOps", "DB-001", "3 days", "Not started"],
    ]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"
    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=INK)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    widths = [13, 19, 18, 36, 54, 10, 14, 22, 12, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(color="0000FF" if cell.column in (7, 9, 10) else INK, size=9)
        row[5].fill = PatternFill("solid", fgColor=YELLOW if row[5].value == "P0" else LIGHT)
        row[9].fill = PatternFill("solid", fgColor=PINK if row[9].value == "Deferred" else MINT)
    dv_priority = DataValidation(type="list", formula1='"P0,P1,P2"')
    dv_status = DataValidation(type="list", formula1='"Not started,In progress,Blocked,Done,Deferred"')
    ws.add_data_validation(dv_priority); dv_priority.add(f"F2:F1000")
    ws.add_data_validation(dv_status); dv_status.add(f"J2:J1000")
    wb.save(path)


def verify_workbook(path: Path, expected_sheets: list[str]) -> None:
    wb = load_workbook(path, data_only=False)
    assert wb.sheetnames == expected_sheets, (path, wb.sheetnames)
    minimum_size = 10_000 if "product-import" in path.name else 6_000
    assert path.stat().st_size > minimum_size, (path, path.stat().st_size)
    if "Products" in wb.sheetnames:
        ws = wb["Products"]
        actual = [str(ws.cell(1, index).value) for index in range(1, len(HEADERS) + 1)]
        assert actual == HEADERS
        assert ws.freeze_panes == "A2"
        assert len(ws.data_validations.dataValidation) >= 7
    wb.close()


def main() -> None:
    import_path = OUT / "product-import-template.xlsx"
    backlog_path = OUT / "implementation-backlog.xlsx"
    create_import_template(import_path)
    create_backlog(backlog_path)
    verify_workbook(import_path, ["Instructions", "Products", "Lists"])
    verify_workbook(backlog_path, ["Backlog"])
    (PUBLIC / import_path.name).write_bytes(import_path.read_bytes())

    csv_path = OUT / "product-import-example.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(HEADERS)
        writer.writerows(SAMPLE_ROWS)

    print(f"Created and verified: {import_path}")
    print(f"Created and verified: {backlog_path}")
    print(f"Created: {csv_path}")


if __name__ == "__main__":
    main()
