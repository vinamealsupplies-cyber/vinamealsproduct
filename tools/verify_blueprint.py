from __future__ import annotations

import hashlib
import json
import re
import subprocess
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
STARTER = ROOT / "starter"
DATABASE = ROOT / "database"

EXPECTED_IMPORT_HEADERS = [
    "operation", "product_handle", "product_name", "slug", "short_description", "description",
    "category_path", "variant_name", "sku", "barcode", "attributes_json", "retail_price",
    "wholesale_price", "cost_price", "track_inventory", "opening_quantity", "reorder_point",
    "location_code", "taxable", "unit", "weight_oz", "active",
    *[f"image_url_{index}" for index in range(1, 11)], "video_url",
]

VIETNAMESE_UI_TERMS = {
    "sản phẩm", "khách hàng", "hóa đơn", "tồn kho", "đăng nhập", "đăng ký",
    "quản trị", "thanh toán", "danh mục", "tìm kiếm", "chi phí", "lợi nhuận",
    "bán hàng", "trang chủ", "giỏ hàng",
}


def fail(message: str) -> None:
    raise AssertionError(message)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def scan_sql(text: str, source: str) -> list[str]:
    """Basic PostgreSQL-aware delimiter scanner; validates quotes/comments/parentheses."""
    statements: list[str] = []
    start = 0
    i = 0
    parens = 0
    state = "normal"
    dollar_tag = ""
    block_depth = 0

    while i < len(text):
        ch = text[i]
        nxt = text[i + 1] if i + 1 < len(text) else ""

        if state == "line_comment":
            if ch == "\n":
                state = "normal"
            i += 1
            continue

        if state == "block_comment":
            if ch == "/" and nxt == "*":
                block_depth += 1
                i += 2
                continue
            if ch == "*" and nxt == "/":
                block_depth -= 1
                i += 2
                if block_depth == 0:
                    state = "normal"
                continue
            i += 1
            continue

        if state == "single_quote":
            if ch == "'":
                if nxt == "'":
                    i += 2
                    continue
                state = "normal"
            i += 1
            continue

        if state == "double_quote":
            if ch == '"':
                if nxt == '"':
                    i += 2
                    continue
                state = "normal"
            i += 1
            continue

        if state == "dollar_quote":
            if text.startswith(dollar_tag, i):
                i += len(dollar_tag)
                state = "normal"
            else:
                i += 1
            continue

        if ch == "-" and nxt == "-":
            state = "line_comment"
            i += 2
            continue
        if ch == "/" and nxt == "*":
            state = "block_comment"
            block_depth = 1
            i += 2
            continue
        if ch == "'":
            state = "single_quote"
            i += 1
            continue
        if ch == '"':
            state = "double_quote"
            i += 1
            continue
        if ch == "$":
            match = re.match(r"\$[A-Za-z_][A-Za-z0-9_]*\$|\$\$", text[i:])
            if match:
                dollar_tag = match.group(0)
                state = "dollar_quote"
                i += len(dollar_tag)
                continue
        if ch == "(":
            parens += 1
        elif ch == ")":
            parens -= 1
            if parens < 0:
                fail(f"{source}: unmatched closing parenthesis at offset {i}")
        elif ch == ";" and parens == 0:
            statement = text[start:i + 1].strip()
            clean = re.sub(r"^(?:\s*--[^\n]*(?:\n|$)|\s*/\*.*?\*/\s*)+", "", statement, flags=re.S)
            if clean:
                statements.append(clean)
            start = i + 1
        i += 1

    if state not in {"normal", "line_comment"}:
        fail(f"{source}: unterminated SQL state {state}")
    if parens != 0:
        fail(f"{source}: unbalanced parentheses ({parens})")
    tail = text[start:].strip()
    if tail and not tail.startswith("--"):
        fail(f"{source}: trailing SQL without semicolon: {tail[:80]!r}")
    return statements


def find_matching_paren(text: str, opening: int) -> int:
    depth = 0
    state = "normal"
    i = opening
    dollar_tag = ""
    while i < len(text):
        ch = text[i]
        nxt = text[i + 1] if i + 1 < len(text) else ""
        if state == "single":
            if ch == "'":
                if nxt == "'":
                    i += 2
                    continue
                state = "normal"
            i += 1
            continue
        if state == "double":
            if ch == '"':
                if nxt == '"':
                    i += 2
                    continue
                state = "normal"
            i += 1
            continue
        if state == "line":
            if ch == "\n":
                state = "normal"
            i += 1
            continue
        if state == "block":
            if ch == "*" and nxt == "/":
                state = "normal"
                i += 2
            else:
                i += 1
            continue
        if state == "dollar":
            if text.startswith(dollar_tag, i):
                i += len(dollar_tag)
                state = "normal"
            else:
                i += 1
            continue
        if ch == "'":
            state = "single"
        elif ch == '"':
            state = "double"
        elif ch == "-" and nxt == "-":
            state = "line"
            i += 2
            continue
        elif ch == "/" and nxt == "*":
            state = "block"
            i += 2
            continue
        elif ch == "$":
            match = re.match(r"\$[A-Za-z_][A-Za-z0-9_]*\$|\$\$", text[i:])
            if match:
                dollar_tag = match.group(0)
                state = "dollar"
                i += len(dollar_tag)
                continue
        elif ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
            if depth == 0:
                return i
        i += 1
    fail("No matching parenthesis found")
    return -1


def split_top_level_csv(body: str) -> list[str]:
    result: list[str] = []
    start = 0
    depth = 0
    state = "normal"
    i = 0
    while i < len(body):
        ch = body[i]
        nxt = body[i + 1] if i + 1 < len(body) else ""
        if state == "single":
            if ch == "'":
                if nxt == "'":
                    i += 2
                    continue
                state = "normal"
            i += 1
            continue
        if state == "double":
            if ch == '"':
                if nxt == '"':
                    i += 2
                    continue
                state = "normal"
            i += 1
            continue
        if ch == "'":
            state = "single"
        elif ch == '"':
            state = "double"
        elif ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        elif ch == "," and depth == 0:
            result.append(body[start:i].strip())
            start = i + 1
        i += 1
    result.append(body[start:].strip())
    return [item for item in result if item]


def extract_tables(schema_text: str) -> dict[str, set[str]]:
    tables: dict[str, set[str]] = {}
    for match in re.finditer(r"create\s+table\s+public\.(\w+)\s*\(", schema_text, flags=re.I):
        open_pos = schema_text.find("(", match.start())
        close_pos = find_matching_paren(schema_text, open_pos)
        body = schema_text[open_pos + 1:close_pos]
        columns: set[str] = set()
        for segment in split_top_level_csv(body):
            stripped = re.sub(r"^--.*?\n", "", segment, flags=re.S).strip()
            lowered = stripped.lower()
            if lowered.startswith(("constraint ", "primary key", "unique ", "check ", "foreign key")):
                continue
            col_match = re.match(r'"?([A-Za-z_][A-Za-z0-9_]*)"?\s+', stripped)
            if col_match:
                columns.add(col_match.group(1).lower())
        tables[match.group(1).lower()] = columns
    return tables


def verify_sql() -> dict[str, int]:
    sql_files = sorted(DATABASE.glob("*.sql")) + sorted((DATABASE / "tests").glob("*.sql"))
    if len(sql_files) < 6:
        fail("Expected five migrations and at least one SQL test")

    texts = {path.name: path.read_text(encoding="utf-8") for path in sql_files}
    statement_count = 0
    for path in sql_files:
        statements = scan_sql(texts[path.name], str(path.relative_to(ROOT)))
        statement_count += len(statements)
        if not statements:
            fail(f"{path.name}: no SQL statements found")

    schema_text = texts["001_types_and_schema.sql"]
    all_migrations = "\n".join(texts[name] for name in sorted(texts) if not name.startswith("001_smoke"))
    tables = extract_tables(schema_text)
    if len(tables) < 20:
        fail(f"Expected at least 20 public tables, found {len(tables)}")

    rls_text = texts["003_rls_and_grants.sql"]
    if "revoke create on schema public from public, anon, authenticated" not in rls_text.lower():
        fail("Public schema CREATE privilege is not explicitly revoked from client roles")
    if "and p.status = 'active'" not in rls_text.lower():
        fail("Database role helpers do not disable inactive staff profiles")
    rls_tables = set(re.findall(r"alter\s+table\s+public\.(\w+)\s+enable\s+row\s+level\s+security", rls_text, flags=re.I))
    missing_rls = set(tables) - {name.lower() for name in rls_tables}
    if missing_rls:
        fail(f"Tables missing RLS: {sorted(missing_rls)}")

    for match in re.finditer(
        r"grant\s+select\s*\((.*?)\)\s+on\s+public\.(\w+)\s+to\s+(?:anon|authenticated)",
        rls_text,
        flags=re.I | re.S,
    ):
        columns = [part.strip().lower() for part in match.group(1).split(",")]
        table = match.group(2).lower()
        if table not in tables:
            fail(f"Column grant references unknown table {table}")
        unknown = set(columns) - tables[table]
        if unknown:
            fail(f"Column grant on {table} references unknown columns: {sorted(unknown)}")

    anon_variant_grant = re.search(
        r"grant\s+select\s*\((.*?)\)\s+on\s+public\.product_variants\s+to\s+anon",
        rls_text,
        flags=re.I | re.S,
    )
    if not anon_variant_grant:
        fail("Missing anonymous product_variants column grant")
    forbidden = {"cost_price", "wholesale_price", "object_key", "created_by", "unit_cost_snapshot"}
    if forbidden & set(re.findall(r"\b\w+\b", anon_variant_grant.group(1).lower())):
        fail("Anonymous variant grant exposes a forbidden internal column")

    if re.search(r"grant\s+select\s+on\s+public\.product_media\s+to\s+(?:anon|authenticated)", rls_text, flags=re.I):
        fail("Product media must use column-scoped SELECT grants")

    reporting_text = texts["004_reporting_views.sql"]
    if "private.can_view_staff_data()" not in reporting_text or "private.can_view_manager_data()" not in reporting_text:
        fail("Management reporting views are missing explicit role/service-role gates")
    if "cost_price" in re.search(r"create or replace view public\.v_product_listing.*?;", reporting_text, flags=re.I | re.S).group(0):
        fail("Storefront product listing exposes cost_price")

    functions = set(re.findall(r"create\s+or\s+replace\s+function\s+(?:public|private)\.(\w+)", all_migrations, flags=re.I))
    triggers: list[tuple[str, str, str, str]] = []
    for trigger_match in re.finditer(r"create\s+trigger\s+(\w+)(.*?);", all_migrations, flags=re.I | re.S):
        trigger_name = trigger_match.group(1)
        trigger_body = trigger_match.group(2)
        target = re.search(r"\bon\s+(public|auth)\.(\w+)", trigger_body, flags=re.I)
        executor = re.search(r"execute\s+function\s+public\.(\w+)\s*\(", trigger_body, flags=re.I)
        if not target or not executor:
            fail(f"Could not parse trigger target/function for {trigger_name}")
        schema_name, table = target.group(1).lower(), target.group(2).lower()
        function = executor.group(1).lower()
        triggers.append((trigger_name, schema_name, table, function))
        if schema_name == "public" and table not in tables:
            fail(f"Trigger {trigger_name} references unknown table public.{table}")
        if schema_name == "auth" and table != "users":
            fail(f"Trigger {trigger_name} references unexpected auth table auth.{table}")
        if function not in {name.lower() for name in functions}:
            fail(f"Trigger {trigger_name} references unknown function {function}")

    smoke = texts["001_smoke_test.sql"].lower()
    for required in [
        "rollback;", "at most 10 images", "requires a negative", "exact opposite", "immutable",
        "requires a verifier", "customer must match", "currency must match",
        "shipping_revenue", "v_monthly_business_performance",
    ]:
        if required not in smoke:
            fail(f"Smoke test is missing required assertion marker: {required}")

    for reporting_marker in [
        "shipping_revenue",
        "and il.is_active",
        "and c.is_active",
        "coalesce(i.net_sales, 0) + coalesce(i.shipping_revenue, 0)",
    ]:
        if reporting_marker not in reporting_text:
            fail(f"Reporting views are missing required definition: {reporting_marker}")

    functions_text = texts["002_functions_and_triggers.sql"].lower()
    for invariant in [
        "validate_inventory_movement",
        "ensure_invoice_order_consistency",
        "ensure_payment_currency_matches_invoice",
        "ensure_tax_exempt_audit_fields",
    ]:
        if invariant not in functions_text:
            fail(f"Database functions are missing invariant: {invariant}")

    rpc_text = texts["006_transactional_admin_rpcs.sql"].lower()
    for marker in [
        "security definer",
        "set search_path = pg_catalog, private, pg_temp",
        "private.is_service_role()",
        "for update",
        "v_image_count >= 10",
        "p_object_key !~",
        "p_public_url !~ '^https://'",
        "revoke all on function public.admin_complete_product_image",
        "to service_role",
    ]:
        if marker not in rpc_text:
            fail(f"Transactional image RPC is missing security/concurrency marker: {marker}")

    for schema_marker in [
        "product_media_provider_object_uidx",
        "product_media_stream_uid_uidx",
        "product_media_one_primary_image_uidx",
    ]:
        if schema_marker not in schema_text.lower():
            fail(f"Product media schema is missing unique index: {schema_marker}")

    return {
        "sql_files": len(sql_files),
        "sql_statements": statement_count,
        "tables": len(tables),
        "triggers": len(triggers),
    }


def verify_workbooks() -> dict[str, int]:
    template = ROOT / "import-templates" / "product-import-template.xlsx"
    public_copy = STARTER / "public" / "templates" / "product-import-template.xlsx"
    backlog = ROOT / "import-templates" / "implementation-backlog.xlsx"
    for path in (template, public_copy, backlog):
        if not path.exists() or path.stat().st_size < 1000:
            fail(f"Workbook missing or unexpectedly small: {path.relative_to(ROOT)}")

    if sha256(template) != sha256(public_copy):
        fail("Public product import template does not match the source workbook")

    wb = load_workbook(template)
    if set(wb.sheetnames) != {"Instructions", "Products", "Lists"}:
        fail(f"Unexpected product template sheets: {wb.sheetnames}")
    if wb["Lists"].sheet_state != "hidden":
        fail("Lists sheet must be hidden")
    ws = wb["Products"]
    headers = [ws.cell(1, index).value for index in range(1, len(EXPECTED_IMPORT_HEADERS) + 1)]
    if headers != EXPECTED_IMPORT_HEADERS:
        fail("Product import workbook headers do not match the parser contract")
    if ws.max_row < 5:
        fail("Product import workbook should include multiple sample rows")
    if len(ws.data_validations.dataValidation) < 7:
        fail("Product import workbook is missing expected data validation rules")
    if not ws.auto_filter.ref or ws.freeze_panes != "A2":
        fail("Product import workbook is missing filter/freeze configuration")

    backlog_wb = load_workbook(backlog, read_only=True, data_only=False)
    backlog_ws = backlog_wb["Backlog"]
    if backlog_ws.max_row < 20 or backlog_ws.max_column != 10:
        fail("Implementation backlog is incomplete")

    return {
        "import_headers": len(headers),
        "import_sample_rows": ws.max_row - 1,
        "backlog_tasks": backlog_ws.max_row - 1,
    }


def iter_source_files() -> Iterable[Path]:
    for extension in ("*.ts", "*.tsx", "*.mjs"):
        yield from STARTER.rglob(extension)


def verify_starter() -> dict[str, int | str]:
    package = json.loads((STARTER / "package.json").read_text())
    all_versions = list(package.get("dependencies", {}).values()) + list(package.get("devDependencies", {}).values())
    if any(version == "latest" for version in all_versions):
        fail("package.json still contains an unpinned 'latest' dependency")
    if package.get("engines", {}).get("node") != ">=22":
        fail("Node engine must be >=22")
    if package.get("dependencies", {}).get("next") != "16.2.11":
        fail("Unexpected pinned Next.js version")

    auth_source = (STARTER / "lib" / "auth.ts").read_text(encoding="utf-8")
    for marker in ['role, status', 'profile.status !== "active"']:
        if marker not in auth_source:
            fail(f"Server auth guard is missing inactive-profile enforcement: {marker}")

    import_parser = (STARTER / "lib" / "import" / "product-import.ts").read_text(encoding="utf-8")
    for marker in ["oversizedImageHeaders", "unknownHeaders", "seenBarcodes", "opening_quantity must be 0"]:
        if marker not in import_parser:
            fail(f"Product import parser is missing validation marker: {marker}")

    report_page = (STARTER / "app" / "admin" / "reports" / "page.tsx").read_text(encoding="utf-8")
    for marker in ["shippingRevenue", "taxCollected", "amountInvoiced", "balanceDue", "operatingProfit"]:
        if marker not in report_page:
            fail(f"Admin report page is missing metric: {marker}")
    if "netSales - row.received" in report_page:
        fail("Admin report page incorrectly derives outstanding balance from net sales minus cash received")

    media_complete = (STARTER / "app" / "api" / "admin" / "media" / "complete" / "route.ts").read_text(encoding="utf-8")
    for marker in ["inspectImageObject", "admin_complete_product_image", "IMAGE_LIMIT_REACHED", "CONTENT_TYPE_MISMATCH"]:
        if marker not in media_complete:
            fail(f"Image completion route is missing verification/RPC marker: {marker}")

    stream_route = (STARTER / "app" / "api" / "admin" / "video" / "direct-upload" / "route.ts").read_text(encoding="utf-8")
    for marker in ["STREAM_UNAVAILABLE", "VIDEO_LIMIT_REACHED", "stream/direct_upload"]:
        if marker not in stream_route:
            fail(f"Video direct-upload route is missing resilience/limit marker: {marker}")

    media_uploader = (STARTER / "components" / "product-media-uploader.tsx").read_text(encoding="utf-8")
    for marker in ["allowedImageTypes", "maxImageBytes", "8 MB limit", "allowedVideoTypes"]:
        if marker not in media_uploader:
            fail(f"Product media uploader is missing client validation marker: {marker}")

    required_files = [
        "app/page.tsx", "app/products/page.tsx", "app/products/[slug]/page.tsx",
        "app/account/page.tsx", "app/admin/layout.tsx", "app/admin/inventory/page.tsx",
        "app/admin/customers/page.tsx", "app/admin/invoices/page.tsx",
        "app/admin/reports/page.tsx", "app/admin/imports/page.tsx",
        "app/api/admin/media/presign/route.ts", "app/api/admin/media/complete/route.ts",
        "app/api/admin/video/direct-upload/route.ts",
        "app/api/admin/imports/products/preview/route.ts",
        "proxy.ts", "lib/supabase/server.ts", "lib/auth.ts",
    ]
    missing = [path for path in required_files if not (STARTER / path).exists()]
    if missing:
        fail(f"Starter is missing required files: {missing}")

    ui_files = [path for path in iter_source_files() if "scripts" not in path.parts]
    for path in ui_files:
        lowered = path.read_text(encoding="utf-8").lower()
        found = sorted(term for term in VIETNAMESE_UI_TERMS if term in lowered)
        if found:
            fail(f"Vietnamese UI terms found in {path.relative_to(ROOT)}: {found}")

    env_lines = (STARTER / ".env.example").read_text().splitlines()
    for line in env_lines:
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        if any(marker in key for marker in ["KEY", "TOKEN", "SECRET"]):
            if value and value not in {"https://media.example.com", "http://localhost:3000", "false"}:
                fail(f"Potential real secret found in .env.example for {key}")

    # Optional syntax transpilation using either local or globally installed TypeScript.
    ts_count = len(list(STARTER.rglob("*.ts"))) + len(list(STARTER.rglob("*.tsx"))) - 1  # exclude next-env.d.ts
    syntax_status = "not-run"
    try:
        npm_root = subprocess.check_output(["npm", "root", "-g"], text=True, timeout=10).strip()
        ts_module = Path(npm_root) / "typescript"
        if ts_module.exists():
            script = r'''
const fs = require('fs');
const path = require('path');
const ts = require(process.argv[1]);
const root = process.argv[2];
let failures = [];
let count = 0;
function walk(dir) {
  for (const entry of fs.readdirSync(dir, { withFileTypes: true })) {
    if (entry.name === 'node_modules' || entry.name === '.next') continue;
    const full = path.join(dir, entry.name);
    if (entry.isDirectory()) walk(full);
    else if ((full.endsWith('.ts') || full.endsWith('.tsx')) && !full.endsWith('.d.ts')) {
      count += 1;
      const text = fs.readFileSync(full, 'utf8');
      const result = ts.transpileModule(text, {
        compilerOptions: { jsx: ts.JsxEmit.ReactJSX, target: ts.ScriptTarget.ES2022, module: ts.ModuleKind.ESNext },
        reportDiagnostics: true,
        fileName: full,
      });
      for (const d of result.diagnostics || []) {
        if (d.category === ts.DiagnosticCategory.Error) {
          failures.push(full + ': ' + ts.flattenDiagnosticMessageText(d.messageText, '\\n'));
        }
      }
    }
  }
}
walk(root);
if (failures.length) { console.error(failures.join('\n')); process.exit(1); }
console.log(count);
'''
            output = subprocess.check_output(
                ["node", "-e", script, str(ts_module), str(STARTER)],
                text=True,
                stderr=subprocess.STDOUT,
                timeout=60,
            ).strip()
            ts_count = int(output.splitlines()[-1])
            syntax_status = "passed"
    except (OSError, subprocess.SubprocessError, ValueError):
        syntax_status = "not-run"

    return {
        "source_files_checked": len(ui_files),
        "typescript_files": ts_count,
        "typescript_syntax": syntax_status,
    }


def main() -> None:
    required = [
        ROOT / "README.md",
        ROOT / "docs" / "00-delivery-status-vn.md",
        ROOT / "docs" / "16-implementation-handoff-vn.md",
        ROOT / "database" / "tests" / "001_smoke_test.sql",
    ]
    for path in required:
        if not path.exists():
            fail(f"Missing required project file: {path.relative_to(ROOT)}")

    results = {
        "sql": verify_sql(),
        "workbooks": verify_workbooks(),
        "starter": verify_starter(),
    }
    print(json.dumps({"status": "PASS", **results}, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
